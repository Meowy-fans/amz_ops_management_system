# @retire(since="2026-06-15", replaced_by="AmazonListingSubmitter strict validation feedback", scheduled_removal="2026-07-31", scope="file")
"""Helpers for correcting Amazon template required-field rules."""
import logging
import os
import re
from typing import Dict, Set, Tuple

import openpyxl

logger = logging.getLogger(__name__)


def parse_report_for_required_fields(file_path: str) -> Set[str]:
    """Parse an Amazon report and return fields missing required values."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"指定的报错文件路径不存在: {file_path}")

    required_fields = set()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if "Feed Processing Summary" not in wb.sheetnames:
        raise ValueError(
            "文件中未找到名为 'Feed Processing Summary' 的工作表。"
        )

    sheet = wb["Feed Processing Summary"]
    header = []
    header_row_idx = -1

    for i in range(1, 10):
        row_values = [
            str(cell.value).strip()
            for cell in sheet[i]
            if cell.value
        ]
        if "Error code" in row_values and "Error message" in row_values:
            header = [
                str(cell.value).strip() if cell.value else ""
                for cell in sheet[i]
            ]
            header_row_idx = i
            break

    if header_row_idx == -1:
        raise ValueError(
            "未能在 'Feed Processing Summary' 中找到包含 "
            "'Error code' 和 'Error message' 的表头。"
        )

    try:
        code_col_idx = header.index("Error code")
        msg_col_idx = header.index("Error message")
    except ValueError:
        raise ValueError(
            "表头中必须同时包含 'Error code' 和 'Error message' 列。"
        )

    pattern = re.compile(r"'(.+?)' is required but not supplied\.")

    for row in sheet.iter_rows(min_row=header_row_idx + 1):
        error_code = row[code_col_idx].value
        error_message = row[msg_col_idx].value

        if str(error_code).strip() == "90220" and isinstance(error_message, str):
            match = pattern.search(error_message)
            if match:
                required_fields.add(match.group(1))

    return required_fields


def apply_required_field_corrections(
    definitions: Dict,
    fields_to_correct: Set[str],
) -> Tuple[Dict, Set[str]]:
    """Mark required-child and required-single fields as required."""
    corrected_fields = set()
    label_to_key_map = {
        v.get("local_label"): k
        for k, v in definitions.items()
        if v and isinstance(v, dict) and v.get("local_label")
    }

    for field_name in fields_to_correct:
        target_key = label_to_key_map.get(field_name)

        if target_key:
            field_def = definitions.get(target_key, {})

            if (
                field_def.get("required_child") != "Required"
                or field_def.get("required_single") != "Required"
            ):
                field_def["required_child"] = "Required"
                field_def["required_single"] = "Required"
                corrected_fields.add(field_name)
        else:
            logger.warning(
                f"在数据库的字段定义中，未能找到 local_label 为 "
                f"'{field_name}' 的记录，已跳过。"
            )

    return definitions, corrected_fields
