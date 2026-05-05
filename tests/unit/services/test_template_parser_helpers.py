import openpyxl

from src.services.amz_template_parser import AdvancedTemplateParser
from src.services.template_parser_helpers import (
    build_data_definition_column_mapping,
    extract_field_definitions,
    extract_valid_values,
    find_data_definition_header,
    is_deprecated,
    parse_valid_value_declaration,
)


def test_data_definition_helpers_find_header_and_extract_grouped_fields():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data Definitions"
    sheet.append(["ignored"])
    sheet.append([
        "Group Name",
        "Field Name",
        "Local Label Name",
        "Accepted Values",
        "Required for Parant?",
    ])
    sheet.append(["Identity", "", "", "", ""])
    sheet.append(["", "brand_name", "Brand", "Free text", "Required"])

    header_row_idx, raw_headers = find_data_definition_header(sheet)
    column_mapping = build_data_definition_column_mapping(raw_headers)
    definitions = extract_field_definitions(sheet, header_row_idx, column_mapping)

    assert header_row_idx == 2
    assert column_mapping["required_parent"] == 4
    assert definitions["brand_name"]["group"] == "Identity"
    assert definitions["brand_name"]["local_label"] == "Brand"
    assert definitions["brand_name"]["required_parent"] == "Required"


def test_valid_value_helpers_parse_scope_and_filter_deprecated_values():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Valid Values"
    sheet.append(["Variation", "", "", ""])
    sheet.append(["", "Variation Theme Name [Global]", "Color", "Old Deprecated"])
    sheet.append(["", "Color Name [Global]", "Red", "Do not use Blue"])

    values = extract_valid_values(sheet, skip_deprecated=True)

    assert parse_valid_value_declaration("Color Name - [Global]") == (
        "Color Name",
        "Global",
    )
    assert is_deprecated("obsolete shade")
    assert values == [
        {
            "group": "Variation",
            "attribute": "Variation Theme Name",
            "scope": "Global",
            "values": ["Color"],
        },
        {
            "group": "Variation",
            "attribute": "Color Name",
            "scope": "Global",
            "values": ["Red"],
        },
    ]


def test_advanced_template_parser_parse_minimal_workbook(tmp_path):
    file_path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()

    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=4, column=1, value="sku")
    template_sheet.cell(row=4, column=2, value="brand_name")

    definitions_sheet = wb.create_sheet("Data Definitions")
    definitions_sheet.append([
        "Group Name",
        "Field Name",
        "Local Label Name",
        "Accepted Values",
        "Required for Parent?",
        "Required for Child?",
    ])
    definitions_sheet.append(["Identity", "", "", "", "", ""])
    definitions_sheet.append(["", "brand_name", "Brand", "Free text", "Required", "Required"])

    valid_values_sheet = wb.create_sheet("Valid Values")
    valid_values_sheet.append(["Variation", "", "", ""])
    valid_values_sheet.append(["", "Variation Theme Name [Global]", "Color", "Size"])

    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is True
    assert message == "解析成功"
    assert parser.get_results()["fields"] == ["sku", "brand_name"]
    assert parser.get_results()["field_definitions"]["brand_name"]["group"] == "Identity"
    assert parser.get_all_variation_themes() == ["Color", "Size"]


def test_advanced_template_parser_open_workbook_failure(tmp_path):
    parser = AdvancedTemplateParser(str(tmp_path / "missing.xlsx"))

    success, message = parser.parse()

    assert success is False
    assert message == "无法打开Excel工作簿"


def test_advanced_template_parser_missing_template_sheet(tmp_path):
    file_path = tmp_path / "missing-template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Data Definitions"
    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is False
    assert message == "解析 'Template' 工作表失败"


def test_advanced_template_parser_empty_template_sheet(tmp_path):
    file_path = tmp_path / "empty-template.xlsx"
    wb = openpyxl.Workbook()
    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=1, column=1, value=" ")
    wb.create_sheet("Data Definitions")
    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is False
    assert message == "解析 'Template' 工作表失败"


def test_advanced_template_parser_missing_data_definitions_sheet(tmp_path):
    file_path = tmp_path / "missing-definitions.xlsx"
    wb = openpyxl.Workbook()
    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=1, column=1, value="sku")
    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is False
    assert message == "解析 'Data Definitions' 工作表失败"


def test_advanced_template_parser_data_definitions_without_header(tmp_path):
    file_path = tmp_path / "bad-definitions.xlsx"
    wb = openpyxl.Workbook()
    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=1, column=1, value="sku")
    definitions_sheet = wb.create_sheet("Data Definitions")
    definitions_sheet.append(["Group", "Label"])
    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is False
    assert message == "解析 'Data Definitions' 工作表失败"


def test_advanced_template_parser_data_definitions_without_field_name(tmp_path):
    wb = openpyxl.Workbook()
    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=1, column=1, value="sku")
    definitions_sheet = wb.create_sheet("Data Definitions")
    definitions_sheet.append(["Group Name", "Local Label Name", "Accepted Values"])

    parser = AdvancedTemplateParser("template.xlsx")
    parser.wb = wb

    assert parser._parse_data_definitions_advanced() is False


def test_advanced_template_parser_valid_values_sheet_is_optional(tmp_path):
    file_path = tmp_path / "without-valid-values.xlsx"
    wb = openpyxl.Workbook()
    template_sheet = wb.active
    template_sheet.title = "Template"
    template_sheet.cell(row=1, column=1, value="sku")
    definitions_sheet = wb.create_sheet("Data Definitions")
    definitions_sheet.append(["Group Name", "Field Name", "Local Label Name"])
    definitions_sheet.append(["Identity", "sku", "SKU"])
    wb.save(file_path)

    parser = AdvancedTemplateParser(str(file_path))
    success, message = parser.parse()

    assert success is True
    assert message == "解析成功"
    assert parser.valid_values == []


def test_advanced_template_parser_handles_parser_exceptions(monkeypatch, tmp_path):
    file_path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Template"
    wb.save(file_path)
    parser = AdvancedTemplateParser(str(file_path))
    monkeypatch.setattr(parser, "_parse_template_sheet", lambda: (_ for _ in ()).throw(RuntimeError("boom")))

    success, message = parser.parse()

    assert success is False
    assert message == "解析失败: boom"


def test_advanced_template_parser_variation_theme_fallback_and_deprecation():
    parser = AdvancedTemplateParser("template.xlsx", skip_deprecated=True)
    parser.valid_values = [{"attribute": "Color Name", "values": ["Red"]}]

    assert parser.get_all_variation_themes() == []
    assert parser._is_deprecated("Old deprecated value") is True
    parser._log_and_print("debug message", "debug")
