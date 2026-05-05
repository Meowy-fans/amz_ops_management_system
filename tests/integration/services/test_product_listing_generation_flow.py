import json
from pathlib import Path
from uuid import UUID

import openpyxl

from src.services.product_listing_service import ProductListingService
from src.utils.data_mapping_helper import DataMappingHelper
from src.utils.excel_generator import ExcelGenerator
from src.utils.variation_helper import VariationHelper


class FakeDbSession:
    def __init__(self):
        self.commits = 0
        self.rollbacks = 0

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1


class FakeProductListingRepository:
    def get_pending_listing_skus(self):
        return ["MEOW-SINGLE", "MEOW-WHITE", "MEOW-BLACK", "MEOW-OTHER"]

    def get_sku_to_category_mapping(self, meow_skus):
        return [
            ("MEOW-SINGLE", "CABINET"),
            ("MEOW-WHITE", "CABINET"),
            ("MEOW-BLACK", "CABINET"),
            ("MEOW-OTHER", "HOME_MIRROR"),
        ]

    def get_variation_data(self, meow_skus):
        return [
            ("MEOW-SINGLE", "GIGA-SINGLE", []),
            ("MEOW-WHITE", "GIGA-WHITE", ["GIGA-BLACK"]),
            ("MEOW-BLACK", "GIGA-BLACK", ["GIGA-WHITE"]),
        ]


class FakeProductDataRepository:
    PRODUCTS = {
        "MEOW-SINGLE": {
            "meow_sku": "MEOW-SINGLE",
            "product_name": "Standalone Cabinet - Oak",
            "category_name": "CABINET",
            "selling_point_1": "Solid wood frame",
            "selling_point_2": "Soft-close doors",
            "raw_data": {"color": "Oak"},
        },
        "MEOW-WHITE": {
            "meow_sku": "MEOW-WHITE",
            "product_name": "Modern Cabinet - White",
            "category_name": "CABINET",
            "selling_point_1": "Water resistant",
            "selling_point_2": "Easy to install",
            "raw_data": {"color": "White"},
        },
        "MEOW-BLACK": {
            "meow_sku": "MEOW-BLACK",
            "product_name": "Modern Cabinet - Black",
            "category_name": "CABINET",
            "selling_point_1": "Water resistant",
            "selling_point_2": "Easy to install",
            "raw_data": {"color": "Black"},
        },
    }

    def get_full_product_data(self, meow_sku):
        return self.PRODUCTS.get(meow_sku)


class FakeTemplateRepository:
    def find_template_by_category(self, category_name):
        assert category_name == "cabinet"
        return {
            "valid_values": [
                {"attribute": "Variation Theme Name", "values": ["Color"]},
                {"attribute": "Product Type", "values": ["CABINET"]},
            ],
            "variation_mapping": {"color_name": "Color"},
            "priority_themes": ["Color"],
        }


class FakeListingLogRepository:
    def __init__(self):
        self.logs = []

    def bulk_insert_log(self, logs):
        self.logs.extend(logs)


class FakeVariationThemeService:
    def determine_variation_theme(self, family_full_data, valid_themes, priority_themes):
        assert valid_themes == ["Color"]
        assert priority_themes == ["Color"]
        return {
            "variation_theme": "Color",
            "child_attributes": {
                item["meow_sku"]: {"color_name": item["raw_data"]["color"]}
                for item in family_full_data
            },
        }


def _write_mapping_config(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "mappings": {
                    "SKU": {"source_type": "direct", "value": "meow_sku"},
                    "Item Name": {"source_type": "direct", "value": "product_name"},
                    "Product Type": {"source_type": "static", "value": "cabinet"},
                    "Brand Name": {"source_type": "static", "value": "Meow"},
                    "Color": {"source_type": "jsonb", "json_path": "color"},
                    "Bullet Point": {
                        "source_type": "db_field_multiple",
                        "fields": ["selling_point_1", "selling_point_2"],
                    },
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _write_template(template_dir: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = [
        "SKU",
        "Item Name",
        "Product Type",
        "Brand Name",
        "Color",
        "Bullet Point",
        "Bullet Point",
        "Listing Action",
        "Relationship Type",
        "Parentage Level",
        "Parent SKU",
        "Child Relationship Type",
        "Variation Theme Name",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=header)
    wb.save(template_dir / "CABINET.xlsm")


def _build_service(tmp_path):
    mapping_path = tmp_path / "amz_mapping.json"
    template_dir = tmp_path / "template_files"
    output_dir = tmp_path / "output"
    template_dir.mkdir()
    output_dir.mkdir()
    _write_mapping_config(mapping_path)
    _write_template(template_dir)

    db = FakeDbSession()
    log_repo = FakeListingLogRepository()
    service = ProductListingService.__new__(ProductListingService)
    service.product_listing_repo = FakeProductListingRepository()
    service.product_data_repo = FakeProductDataRepository()
    service.template_repo = FakeTemplateRepository()
    service.listing_log_repo = log_repo
    service.data_mapper = DataMappingHelper(config_path=mapping_path)
    service.excel_generator = ExcelGenerator(
        template_base_path=template_dir,
        output_base_path=output_dir,
    )
    service.variation_helper = VariationHelper()
    service.category_config = {}
    service.llm_service = None
    service.variation_theme_service = FakeVariationThemeService()
    service.db = db
    return service, db, log_repo


def test_generate_listing_flow_creates_workbook_and_logs_for_single_and_variation(tmp_path):
    service, db, log_repo = _build_service(tmp_path)

    result = service.generate_listings_by_category("cabinet")

    assert result["success"] is True
    assert isinstance(result["batch_id"], UUID)
    assert result["single_count"] == 1
    assert result["variation_count"] == 1
    assert result["total_rows"] == 4
    assert db.commits == 1
    assert db.rollbacks == 0
    assert Path(result["excel_file"]).exists()

    workbook = openpyxl.load_workbook(result["excel_file"], keep_vba=True)
    rows = list(workbook["Template"].iter_rows(min_row=7, max_row=10, values_only=True))
    row_by_sku = {row[0]: row for row in rows}

    assert row_by_sku["MEOW-SINGLE"][1] == "Standalone Cabinet - Oak"
    assert row_by_sku["MEOW-SINGLE"][2] == "CABINET"
    assert row_by_sku["MEOW-SINGLE"][5] == "Solid wood frame"
    assert row_by_sku["MEOW-SINGLE"][6] == "Soft-close doors"
    assert row_by_sku["MEOW-SINGLE"][7] == "Create or Replace (Full Update)"

    parent_skus = [sku for sku in row_by_sku if str(sku).startswith("PARENT-")]
    assert len(parent_skus) == 1
    parent_row = row_by_sku[parent_skus[0]]
    assert parent_row[8] == "Parent"
    assert parent_row[9] == "Parent"
    assert parent_row[12] == "Color"

    for child_sku, expected_color in {
        "MEOW-WHITE": "White",
        "MEOW-BLACK": "Black",
    }.items():
        child_row = row_by_sku[child_sku]
        assert child_row[4] == expected_color
        assert child_row[8] == "Child"
        assert child_row[9] == "Child"
        assert child_row[10] == parent_skus[0]
        assert child_row[11] == "Variation"
        assert child_row[12] == "Color"

    assert {log["meow_sku"] for log in log_repo.logs} == {
        "MEOW-SINGLE",
        "MEOW-WHITE",
        "MEOW-BLACK",
    }
    assert all(log["listing_batch_id"] == result["batch_id"] for log in log_repo.logs)
    assert {log["status"] for log in log_repo.logs} == {"GENERATED"}
